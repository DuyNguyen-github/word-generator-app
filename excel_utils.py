"""
Utilities để đọc và xử lý Excel files.
"""

import datetime
from openpyxl import load_workbook


def get_sheet_names(excel_path):
    """Lấy danh sách tên sheet từ Excel file."""
    try:
        wb = load_workbook(excel_path)
        return wb.sheetnames
    except Exception as e:
        raise Exception(f"Không thể đọc Excel: {e}")


def read_excel_sheet(excel_path, sheet_name=None, header_row=1, data_start_row=2):
    """
    Đọc sheet Excel và trả về (headers, rows).
    
    Args:
        excel_path: đường dẫn Excel file
        sheet_name: tên sheet (None = sheet active)
        header_row: số hàng header (1-indexed)
        data_start_row: hàng bắt đầu dữ liệu (1-indexed)
    
    Returns:
        headers: danh sách tên cột
        rows: danh sách dict {column_name: value}
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        
        if sheet_name is None:
            ws = wb.active
        else:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại")
            ws = wb[sheet_name]
        
        # Đọc header - lấy từ hàng và dừng ở ô trống đầu tiên
        headers = []
        header_row_obj = ws[header_row]
        for cell in header_row_obj:
            val = cell.value
            if val is None:
                break  # dừng ở cột trống đầu tiên
            headers.append(str(val).strip())
        
        if not headers:
            raise ValueError(f"Hàng {header_row} không có header")
        
        # Đọc dữ liệu - lấy tất cả dòng (không filter empty rows)
        rows = []
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_dict = {}
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                raw = cell.value
                
                if raw is None:
                    display = ""
                elif isinstance(raw, (datetime.datetime, datetime.date)):
                    # Chuyển date/datetime sang dd/mm/YYYY
                    display = raw.strftime("%d/%m/%Y")
                elif isinstance(raw, (int, float)):
                    # Format số với dấu chấm ngăn cách
                    if isinstance(raw, float) and raw.is_integer():
                        display = f"{int(raw):,}".replace(",", ".")
                    else:
                        display = f"{raw:,}".replace(",", ".")
                else:
                    display = str(raw).strip()
                
                row_dict[header] = display
            
            # Thêm ALL rows (including empty ones)
            rows.append(row_dict)
        
        return headers, rows
    
    except Exception as e:
        raise Exception(f"Lỗi khi đọc Excel: {e}")


def get_preview_rows(headers, rows, count=3):
    """Lấy preview từ N dòng đầu tiên."""
    return rows[:count]
